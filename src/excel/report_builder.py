from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.clickhouse.client import get_client

PROJECT_ROOT = Path(__file__).resolve().parents[2]
REPORTS_DIR = PROJECT_ROOT / "reports" / "weekly"
TEMPLATES_DIR = PROJECT_ROOT / "excel_templates"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def fetch_market_summary(days: int = 7) -> pd.DataFrame:
    client = get_client()
    date_to = date.today()
    date_from = date_to - timedelta(days=days)
    query = f"""
    SELECT
        m.country_code,
        m.city,
        sum(s.total_rides) AS total_rides,
        round(sum(s.total_revenue_usd), 2) AS revenue_usd,
        round(avg(s.active_users), 0) AS avg_dau,
        round(avg(s.avg_ride_duration_sec) / 60, 2) AS avg_duration_min,
        sum(s.maintenance_events) AS maintenance_events
    FROM mv_market_daily_summary s
    JOIN dim_market m ON s.market_id = m.market_id
    WHERE s.summary_date BETWEEN toDate('{date_from}') AND toDate('{date_to}')
    GROUP BY m.country_code, m.city
    ORDER BY revenue_usd DESC
    """
    return client.query_df(query)


def fetch_zone_heatmap(market_code: str, days: int = 7) -> pd.DataFrame:
    client = get_client()
    date_to = date.today()
    date_from = date_to - timedelta(days=days)
    query = f"""
    SELECT
        z.zone_name,
        z.zone_type,
        sum(k.ride_count) AS rides,
        round(sum(k.revenue_usd), 2) AS revenue_usd,
        round(avg(k.utilization_rate), 4) AS utilization
    FROM mv_daily_zone_kpi k
    JOIN dim_zone z ON k.zone_id = z.zone_id
    JOIN dim_market m ON k.market_id = m.market_id
    WHERE m.country_code = '{market_code}'
      AND k.kpi_date BETWEEN toDate('{date_from}') AND toDate('{date_to}')
    GROUP BY z.zone_name, z.zone_type
    ORDER BY revenue_usd DESC
    LIMIT 30
    """
    return client.query_df(query)


def fetch_rebalancing(
    market_code: str, target_date: date | None = None, limit: int = 20
) -> pd.DataFrame:
    client = get_client()
    td = target_date or (date.today() - timedelta(days=1))
    query = f"""
    SELECT
        v.zone_name,
        v.zone_type,
        v.center_lat,
        v.center_lon,
        round(v.idle_hours_pct, 4) AS idle_pct,
        v.daily_rides,
        round(v.rebalancing_score, 4) AS rebalancing_score,
        multiIf(
            v.rebalancing_score > 0.5, 'Срочно: вывезти',
            v.rebalancing_score > 0.2, 'Средний приоритет',
            'Низкий приоритет'
        ) AS action
    FROM v_rebalancing_score v
    WHERE v.snapshot_date = toDate('{td}')
      AND v.market_id = (SELECT market_id FROM dim_market WHERE country_code = '{market_code}' LIMIT 1)
    ORDER BY v.rebalancing_score DESC
    LIMIT {limit}
    """
    return client.query_df(query)


def build_weekly_market_report(output_path: Path | None = None) -> Path:
    """Generate Weekly Market Performance Excel workbook."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = output_path or REPORTS_DIR / f"weekly_market_{date.today().isoformat()}.xlsx"

    summary = fetch_market_summary()
    wb = Workbook()
    ws = wb.active
    ws.title = "Market Summary"

    ws["A1"] = "JET Fleet Intelligence — Weekly Market Performance"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Период: последние 7 дней | Сформировано: {date.today()}"

    headers = list(summary.columns)
    start_row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h)
    _style_header(ws, start_row, len(headers))

    for r_idx, row in summary.iterrows():
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=start_row + 1 + r_idx, column=c_idx, value=val)

    # Conditional formatting on revenue
    rev_col = headers.index("revenue_usd") + 1
    data_end = start_row + len(summary)
    col_letter = get_column_letter(rev_col)
    ws.conditional_formatting.add(
        f"{col_letter}{start_row + 1}:{col_letter}{data_end}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )

    # Chart: revenue by market
    chart = BarChart()
    chart.title = "Выручка по рынкам (USD)"
    chart.y_axis.title = "USD"
    data = Reference(ws, min_col=rev_col, min_row=start_row, max_row=data_end)
    cats = Reference(ws, min_col=2, min_row=start_row + 1, max_row=data_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 10
    ws.add_chart(chart, f"A{data_end + 3}")

    # Zone heatmap sheet for UZ (Tashkent scenario)
    ws2 = wb.create_sheet("Zone Heatmap UZ")
    zones = fetch_zone_heatmap("UZ")
    headers2 = list(zones.columns)
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, len(headers2))
    for r_idx, row in zones.iterrows():
        for c_idx, val in enumerate(row, 1):
            ws2.cell(row=2 + r_idx, column=c_idx, value=val)
    ws2.freeze_panes = "A2"

    for ws_sheet in [ws, ws2]:
        for col in ws_sheet.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws_sheet.column_dimensions[get_column_letter(col[0].column)].width = min(
                max_len + 2, 40
            )

    wb.save(output_path)
    return output_path


def build_rebalancing_sheet(market_code: str = "UZ", output_path: Path | None = None) -> Path:
    """Generate Ops Rebalancing Sheet for field teams."""
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = (
        output_path or REPORTS_DIR / f"rebalancing_{market_code}_{date.today().isoformat()}.xlsx"
    )

    df = fetch_rebalancing(market_code)
    wb = Workbook()
    ws = wb.active
    ws.title = "Rebalancing"

    ws["A1"] = f"JET Ops — Rebalancing Priority ({market_code})"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = "Rebalancing Score = (idle_hours_pct × demand_gap) / sqrt(snapshots + 1)"

    headers = list(df.columns)
    start = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=start, column=col, value=h)
    _style_header(ws, start, len(headers))

    for r_idx, row in df.iterrows():
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=start + 1 + r_idx, column=c_idx, value=val)

    score_col = headers.index("rebalancing_score") + 1
    end_row = start + len(df)
    col_letter = get_column_letter(score_col)
    ws.conditional_formatting.add(
        f"{col_letter}{start + 1}:{col_letter}{end_row}",
        ColorScaleRule(start_type="min", start_color="63BE7B", end_type="max", end_color="F8696B"),
    )
    ws.freeze_panes = f"A{start + 1}"

    wb.save(output_path)
    return output_path


def build_plotly_html(days: int = 7) -> Path:
    """Export interactive Plotly chart as HTML."""
    import plotly.express as px

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    summary = fetch_market_summary(days)
    fig = px.bar(
        summary,
        x="city",
        y="revenue_usd",
        color="country_code",
        title="Выручка по рынкам JET (USD, 7 дней)",
        labels={"revenue_usd": "Выручка USD", "city": "Город"},
    )
    out = REPORTS_DIR / f"weekly_chart_{date.today().isoformat()}.html"
    fig.write_html(str(out))
    return out
